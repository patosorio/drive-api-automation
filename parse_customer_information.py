import io
from typing import Any, BinaryIO

import google.auth
import googleapiclient
from googleapiclient.discovery import build  # type: ignore
from openpyxl import load_workbook


def download_excel_file(
    drive_service: Any,
    file_id: str,
) -> BinaryIO | None:
    """Download an Excel file from Google Drive.

    Args:
        drive_service (Any): The Google Drive service object.
        file_id (str): The ID of the file to download.

    Returns:
        Optional[BinaryIO]: A binary file-like object
        containing the file's content.
    """
    try:
        request = drive_service.files().get_media(fileId=file_id)
        file_buffer = io.BytesIO()
        downloader = googleapiclient.http.MediaIoBaseDownload(
            file_buffer,
            request)

        done = False
        while not done:
            status, done = downloader.next_chunk()
            print(f'Download {int(status.progress() * 100)}% complete.')

        file_buffer.seek(0)
        return file_buffer

    except googleapiclient.errors.HttpError as http_error:
        print(f'HTTP error occurred: {http_error}')
    except Exception as error:
        print(f'Error downloading file: {error}')
    return None


def retrieve_customer_information(
    file_content: BinaryIO,
) -> list[dict[str, Any]]:
    """
    Parse the first sheet of the Excel file to retrieve customer information.

    Args:
        file_content (BinaryIO): The binary content of the Excel file.

    Returns:
        list[dict[str, Any]]: A list of dictionaries
        containing customer information.

    Raises:
        ValueError: If the Excel format does not match the expected structure.
    """

    workbook = load_workbook(file_content)
    sheet = workbook.active
    if sheet is None:
        raise ValueError(
            'The workbook does not contain any active sheet.')

    expected_columns = [
        '#',
        'CIF',
        'Fullname',
        'ID_No.',
        'Emboss_Name',
        'CASA_Account_No',
        'CARDNUMBER',
        'ISS_DATE',
    ]

    header = [cell.value for cell in next(
        sheet.iter_rows(min_row=2, max_row=2))]

    for index, expected_column in enumerate(expected_columns):
        if index >= len(header) or header[index] != expected_column:
            raise ValueError(
                f'Invalid Excel format.'
                f'Expected column {expected_column} in position {index + 1} '
                f"but got {header[index] if index < len(header) else 'None'}.",
            )

    ignored_columns = header[len(expected_columns):]
    if ignored_columns:
        print(f'Warning: Ignored extra columns: {ignored_columns}')

    customer_data = []

    # Get indices of mandatory columns in expected_columns
    fullname_idx, id_no_idx = map(
        expected_columns.index, ['Fullname', 'ID_No.'])
    for row in sheet.iter_rows(
        min_row=3,
        max_col=len(expected_columns),
        values_only=True,
    ):
        if not any(row):
            continue
        if row[fullname_idx] and row[id_no_idx]:
            customer_data.append(dict(zip(
                expected_columns,
                row,
                strict=True)))

    return customer_data


def main() -> None:
    """Main function to download and parse the Excel file."""
    scopes = ['https://www.googleapis.com/auth/drive']
    try:
        # Authenticate and build the Drive service
        credentials, _ = google.auth.default(scopes=scopes)
        drive_service = build('drive', 'v3', credentials=credentials)

        file_id = '1qXV6uoHz0fTQCKmEFiutuYXu4_g_UDIl'
        file_content = download_excel_file(drive_service, file_id)

        if file_content:
            try:
                customer_data = retrieve_customer_information(file_content)
                print('Customer Information:')
                for record in customer_data:
                    print(record)
            except ValueError as error:
                print(error)
        else:
            print('File download failed.')

    except Exception as error:
        print(f'Error initializing Drive service: {error}')


if __name__ == '__main__':
    main()
